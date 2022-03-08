#!/usr/bin/env python3
#
#

from operator import itemgetter
import sys
import openpyxl
from openpyxl.styles import Font
from datetime import date
import os

class Spreadsheet():

  def __init__(self, dir, filename):
    self.workbook = openpyxl.Workbook()
    today = date.today().strftime('%Y-%m-%d')
    self.filepath = os.path.join(dir, filename + '-' + today + '.xlsx') #save workbook in format 'name-YYYY-MM-DD.xlsx'
    self.workbook.save(self.filepath) 
    self.workbook['Sheet'] #delete default sheet, we'll add them as we need

  def add_sheet(self, name):
    self.workbook.active = self.workbook.create_sheet(title=name)
    self.workbook.active['A1'] = self.filepath + ': ' + name
    self.workbook.save(self.filepath)

  def add_row(self, data, font=None):
    sheet = self.workbook.active
    newrow = sheet.max_row + 1
    for index, item in enumerate(data):
      sheet.cell(row=newrow, column=index+1).value = item
      if font != None:
        sheet.cell(row=newrow, column=index+1).font = font
    self.workbook.save(self.filepath)
  
  def close_file(self):
    self.workbook.save(self.filepath)
    self.workbook.close()
